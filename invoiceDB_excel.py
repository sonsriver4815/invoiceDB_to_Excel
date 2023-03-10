import csv
import sqlite3
import glob
import os
import pandas as pd
import pathlib
from contextlib import closing
import openpyxl as excel
import unicodedata
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox


#データベース名
dbname = "SQLite\invoice.sqlite3"

# インボイス(csv)フォルダ内のcsvを全て開く
file_list = glob.glob('インボイス(csv)\*\*.csv')
# 法人番号
houzin_code = glob.glob('法人番号\*.csv')
# 顧客リスト
houzin_code = glob.glob('名寄せ.csv')



def excel_to_csv():
    #エクセル読込
    #エクセル読込
    wb = excel.load_workbook('取引先マスタ\取引先情報照会.xlsx')
    ws1 = wb['Sheet1']
    #ws2 = wb['Sheet2']

    wb1 = excel.Workbook()
    w1s1 = wb1['Sheet']

    #置換する文字列を指定
    Henkan_mae = ['㈱','（株）','(株)','㈲','（有）','(有)','　',' ','･']
    Henkan_go = ['株式会社','株式会社','株式会社','有限会社','有限会社','有限会社','','','・']
    post = ['-','']

    i=0
    m=0

    #文字置換
    for list in Henkan_mae:
        i = i + 1
    
        for row in ws1.iter_rows():
            for cell in row:
                if cell.col_idx == 17:
            
                    if list in (cell.value or []):

                        #置換
                        new_text1 = cell.value.replace(list, Henkan_go[i-1])
                        cell.value = new_text1
    #文字置換
    for list in post:
        m = m + 1
    
        for row in ws1.iter_rows():
            for cell in row:
                if cell.col_idx == 25:
            
                    if list in (cell.value or []):

                        #置換
                        new_text2 = cell.value.replace(list,'')
                        cell.value = new_text2
                             
    #コピー開始行
    j = 1
    k = 1
    l = 1
             
    for row in ws1.iter_rows():
            for cell in row:
                if cell.col_idx == 17:
                    w1s1.cell(row=j, column=1).value = cell.value
                    j = j + 1
                
    for row in ws1.iter_rows():
            for cell in row:
                if cell.col_idx == 25:
                    w1s1.cell(row=k, column=2).value = cell.value
                    k = k + 1
                   
    for row in ws1.iter_rows():
            for cell in row:
                if cell.col_idx == 26:
                    w1s1.cell(row=l, column=3).value = cell.value
                    l = l + 1
                

    # 別名保存（読み込んだのと別名を指定
    wb1.save("名寄せ.xlsx")

    data = pd.read_excel('名寄せ.xlsx', 'Sheet',header=0)
    #print(data)

    data.to_csv('名寄せ' + '.csv', encoding='utf_8', index=True, header=True)



def creat_table():
    # SQLite3のデータベースを開く
    conn = sqlite3.connect(dbname)
    c = conn.cursor()

    #既に存在しているテーブルを削除
    c.execute("DROP TABLE IF EXISTS invoice")
    c.execute("DROP TABLE IF EXISTS kaisyaryaku")
    c.execute("DROP TABLE IF EXISTS houzincode")
    c.execute("DROP TABLE IF EXISTS kokyaku")

    c.execute("DROP VIEW IF EXISTS インボイス")
    c.execute("DROP VIEW IF EXISTS 顧客")

    #SQLiteの空き領域の開放
    c.execute("VACUUM")

    # テーブルを作る
    c.execute('''CREATE TABLE IF NOT EXISTS invoice (
        registratedNumber text,
        kind INTEGER,
        country INTEGER,
        address text,
        addressPrefectureCode INTEGER,
        addressCityCode INTEGER,
        addressRequest text,
        addressRequestPrefectureCode INTEGER,
        addressRequestCityCode INTEGER,
        name text)''')

    #法人番号テーブル
    c.execute('''CREATE TABLE IF NOT EXISTS houzincode (
        sequenceNumber text,
        corporateNumber text,
        name text, kind text,
        prefectureName text,
        cityName text,
        streetNumber text,
        prefectureCode text,
        cityCode text,
        postCode text)''')


    c.execute('''CREATE TABLE IF NOT EXISTS kaisyaryaku (
        ryaku text,
        kindname text)''')

    c.execute('''CREATE TABLE IF NOT EXISTS kokyaku (
        Number INTEGER,
        Name text,
        Postcode text,
        Address text)''')


    c.execute('begin')
    c.execute('commit')
    conn.close()



def insert_data():
    # SQLite3のデータベースを開く
    conn = sqlite3.connect(dbname)
    c = conn.cursor()
    c.execute('begin')
    
    #CSVファイルを開く
    for i in file_list:
        with open(i,'rt',encoding="utf_8") as fp:
          # CSVを読み込む
          reader = csv.reader(fp)
          # 一行ずつ処理する
          for row in reader:
              # 必要なフィールドだけ取り出す
              registratedNumber = row[1] # 登録番号
              kind = row[4] # 人格区分
              country  = row[5] # 国内外区分
              address = row[11] # 事務所の所在地（法人） 
              addressPrefectureCode  = row[12] # 都道府県コード（法人）

              addressCityCode = row[13] # 市区町村コード（法人）
              addressRequest = row[14] #事務所の所在地（公表申出）
              addressRequestPrefectureCode = row[15] #地都道府県コード（公表申出)
              addressRequestCityCode = row[16] #市区町村コード（公表申出)
              name = row[18] #名称

              if registratedNumber == '以下に掲載がない場合': registratedNumber = ''
              if name == '以下に掲載がない場合': name = ''
              if address == '以下に掲載がない場合': address = ''
              if addressPrefectureCode == '以下に掲載がない場合': addressPrefectureCode = ''
              if addressCityCode == '以下に掲載がない場合': addressCityCode = ''
              if addressRequest == '以下に掲載がない場合': addressRequest = ''
              if addressRequestPrefectureCode == '以下に掲載がない場合': addressRequestPrefectureCode = ''
              if addressRequestCityCode == '以下に掲載がない場合': addressRequestCityCode = ''
      
              # SQLiteに追加
              c.execute('''INSERT INTO invoice (registratedNumber,kind,country,address,
                        addressPrefectureCode,addressCityCode,addressRequest,
                        addressRequestPrefectureCode,addressRequestCityCode,name)
                VALUES(?,?,?,?,?,?,?,?,?,?)''', (registratedNumber,kind,country,address,addressPrefectureCode,
                                                 addressCityCode,addressRequest,addressRequestPrefectureCode,
                                                 addressRequestCityCode,name))


    #CSVファイルを開く
    with open('法人番号\zenkokuall.csv', 'rt', encoding='"Shift-JIS"') as f:
      # CSVを読み込む
      reader2 = csv.reader(f)
      # 一行ずつ処理する
      for row in reader2:
          sequenceNumber = row[0] # 一連番号
          corporateNumber = row[1] # 法人番号
          name  = row[6] # 商号又は名称
          kind = row[8] # 法人種別
          prefectureName = row[9] # 国内所在地(都道府県)
          cityName = row[10] # 国内所在地(市区町村)
          streetNumber = row[11] # 国内所在地(丁目番地等)
          prefectureCode = row[13] # 都道府県コード
          cityCode = row[14] # 市区町村コード
          postCode = row[15] # 郵便番号
      
          if cityName == '以下に掲載がない場合': cityName = ''
          if streetNumber == '以下に掲載がない場合': streetNumber = ''
      
          # SQLiteに追加
          c.execute('''INSERT INTO houzincode (sequenceNumber,corporateNumber,name,kind,
                    prefectureName,cityName,streetNumber,prefectureCode,cityCode,postCode)
            VALUES(?,?,?,?,?,?,?,?,?,?)''', (sequenceNumber,corporateNumber,name,kind,prefectureName,
                                 cityName,streetNumber,prefectureCode,cityCode,postCode))

    #CSVファイルを開く
    with open('名寄せ.csv', 'rt', encoding='utf_8') as fd:
      # CSVを読み込む
      reader3 = csv.reader(fd)
      # 一行ずつ処理する
      for row in reader3:
          Number = row[0] # 一連番号
          Name  = row[1] # 商号又は名称
          Postcode = row[2] #郵便番号
          Address = row[3] #住所
      
          if Number == '以下に掲載がない場合': Number = ''
          if Name == '以下に掲載がない場合': Name = ''
          if Postcode == '以下に掲載がない場合': Postcode = ''
          if Address == '以下に掲載がない場合': Address = ''
      
          # SQLiteに追加
          c.execute('''INSERT INTO kokyaku (Number,Name,Postcode,Address)
            VALUES(?,?,?,?)''', (Number,Name,Postcode,Address))



    #会社略称データ作成
    ryaku_data = [
         ( '（株）',    '株式会社')
        ,( '(株)',    '株式会社' )
        ,( '㈱',    '株式会社' )
        ,( '（有）',    '有限会社')
        ,( '(有)',    '有限会社' )
        ,( '㈲',    '有限会社' )
        ]
  
          
    #データ挿入
    try:
        c.executemany( 'insert into kaisyaryaku (ryaku, kindname) values (?,?)',  ryaku_data )
    except:
        pass
    finally:
        pass
    
    # データベースを閉じる
    c.execute('commit')
    conn.close()

def SQL_Query():
    # SQLite3のデータベースを開く
    conn = sqlite3.connect(dbname)
    c = conn.cursor()
    c.execute('begin')
    
    # Query for LEFT OUTER JOIN
    sql = '''CREATE VIEW [インボイス] AS 
    SELECT 
    [houzincode].[corporateNumber] AS [corporateNumber],
    [invoice].[registratedNumber] AS [registratedNumber],
    [houzincode].[name] AS [name],
    [houzincode].[kind] AS [kind],
    [houzincode].[prefectureName] AS [prefectureName],
    [houzincode].[cityName] AS [cityName],
    [houzincode].[streetNumber] AS [streetNumber],
    [houzincode].[prefectureCode] AS [prefectureCode],
    [houzincode].[cityCode] AS [cityCode],
    [houzincode].[postCode] AS [postCode],
    [invoice].[address] AS [address],
    [invoice].[addressPrefectureCode] AS [addressPrefectureCode],
    [invoice].[addressCityCode] AS [addressCityCode]
    FROM
    [houzincode]
    LEFT OUTER JOIN [invoice]
    ON [invoice].[name] = [houzincode].[name]
    AND [invoice].[addressPrefectureCode] = [houzincode].[prefectureCode]
    AND [invoice].[addressCityCode] = [houzincode].[cityCode]'''


    sql1 = '''CREATE VIEW [顧客] AS 
    SELECT 
    [kokyaku].[Number] AS [Number],
    [kokyaku].[Name] AS [Name],
    [kokyaku].[Postcode] AS [Postcode],
    [kokyaku].[Address] AS [Address],
    [インボイス].[corporateNumber] AS [corporateNumber],
    [インボイス].[registratedNumber] AS [registratedNumber],
    [インボイス].[prefectureName] AS [prefectureName],
    [インボイス].[cityName] AS [cityName],
    [インボイス].[streetNumber] AS [streetNumber],
    [インボイス].[address] AS [address]
    FROM
    [kokyaku]
    LEFT OUTER JOIN [インボイス]
    ON [インボイス].[name] = [kokyaku].[Name];'''

    c.execute(sql)
    c.execute(sql1)
    
    # データベースを閉じる
    c.execute('commit')
    conn.close()

def output_csv():
    #データベース名
    dbname = "SQLite\invoice.sqlite3"

    # SQLite3のデータベースを開く
    conn = sqlite3.connect(dbname)
    c = conn.cursor()
    c.execute('begin')

    # dbをpandasで読み出す。
    df = pd.read_sql('SELECT * FROM 顧客', conn)

    df.to_csv("取引先インボイス連携.csv")

    c.execute('commit')
    conn.close()

def output_excel():
    # CSVファイルの読み込み
    data = pd.read_csv('取引先インボイス連携.csv')
 
    # Excel形式で出力
    data.to_excel('取引先インボイス連携.xlsx', encoding='utf-8')

def deleet_file():
    os.remove('名寄せ.csv')
    os.remove('名寄せ.xlsx')
    os.remove('取引先インボイス連携.csv')
    os.remove('取引先インボイス連携.csv')

#excel_to_csv()
#creat_table()
#insert_data()
#SQL_Query()

#output_csv()
#output_excel()


#GUI作成
class Application(tk.Frame):
    def __init__(self, master = None):
        super().__init__(master)

        self.master.title("取引先インボイス連携")     # ウィンドウタイトル
        self.master.geometry("350x150")       # ウィンドウサイズ(幅x高さ)


        button1 = tk.Button(self.master, text = "①取引先データ変換", 
                            command = self.button_click 
                            )
        
        button2 = tk.Button(self.master, text = "②インボイスデータ作成", 
                            command = self.button_click1 
                            )
        button3 = tk.Button(self.master, text = "③エクセル出力", 
                            command = self.button_click2 
                            )

        # ボタンの配置
        button1.pack(pady=10)
        button2.pack(pady=10)
        button3.pack(pady=10)


    def button_click(self):
        excel_to_csv()
        messagebox.showinfo('①取引先データ変換', '完了しました')
        

    def button_click1(self):
        creat_table()
        insert_data()
        SQL_Query()
        messagebox.showinfo('②インボイスデータ作成', '完了しました')
        
    def button_click2(self):
        output_csv()
        output_excel()
        messagebox.showinfo('③エクセル出力', '完了しました')
        deleet_file()

if __name__ == "__main__":
    root = tk.Tk()
    app = Application(master = root)
    app.mainloop()

